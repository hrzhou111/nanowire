# -*- coding: utf-8 -*-
"""
Created on Wed Oct 18 18:45:09 2017

@author: zhr
"""

import openpyxl as opx
import gdspy
import numpy
import shape

def nanowire(number,width,pitch,length,n,is_interrupt,device_shape,pals_length,itr,elec_width,pin_length=60):
    """
    用于绘制任意大小，并联根数的纳米线版图，python版本：2.7.14,3.3
    长度单位：μm
    width 单根纳米线宽度
    length 纳米线长度，同时是纳米线区域的半径  
    pitch 纳米线周期宽度
    itr 计数纳米线根数
    n:并联纳米线根数
    每隔30um并联纳米线
    elec_width MA6光刻电极宽度
    绘制纳米线直径<=50μm的纳米线探测器
    """
    length = float(length)
    width = float(width)
    pitch = float(pitch)
    n = int(n)
    is_interrupt = int(is_interrupt)
    device_shape = int(device_shape)
    pals_length = float(pals_length)
    elec_width = int(elec_width)
    pin_length = float(pin_length)
    
    minedge = 10
    maxdiameter = length+2*minedge #外围结构为5um，设置纳米线电子束曝光尺寸大小
    Cellname = "nanowire"+str(itr)
    single_cell = gdspy.Cell(Cellname,exclude_from_current=True)
    mark = itr
    #gdspy.GdsLibrary.add(single_cell,overwrite_duplication = True)

    if is_interrupt==1 and pals_length < pitch-width:
        print("并联长度过小")
        return single_cell
    if length+minedge>maxdiameter:
        print("纳米线长度设置应小于等于50μm")
        return single_cell    #返回空的cell
    if is_interrupt==1 and pals_length>length:
        print("纳米线并联长度应小于直径")
        return single_cell    #返回空的cell

    total = int(length/pitch);
    """
    #根据并联纳米线根数和纳米线间距来微调纳米线直径
    """
    if n<1:
        print("纳米线小于1，n应该>=1")
        return single_cell
    elif type(n)==float:
        print('纳米线并联数应为>=1的整数')
        return single_cell
    unnece = total%n   #多余的纳米线根数
    itr = int(total/n)     #纳米线周期
    if unnece>int(n/2):
        total = (itr+1)*n   
    else:
        total = itr*n


    a = (elec_width-length)/pin_length/2
    if a-int(a)>0.8:
        a = int(a)+2
    else:
        a = int(a)+1
    pitch1 = pitch
    if length>pin_length:   #微调周期以保证拼接地方在曝光区域
        pitch = float((pin_length-0.01)/int((pin_length-0.01)/pitch))

    oldl = length
    print("对纳米线"+Cellname+"长度进行了微调：pitch origin = {},pitch={},原始值{}，微调后值{},".format(pitch1,pitch,length,total*pitch))
#   length = total*pitch
#    total1 = total
    half = total/2

    """
    绘制纳米线核心区域
    两步绘制：
    计算调整圆的直径,确定圆的刻蚀线终点
    1 绘制中心圆区域
    2，绘制隔离线
    """
    #步骤一1
    if device_shape==0:
        edge = shape.circle(pitch*total,width,pitch,total,half)
#        iso = shape.circleiso(length,width,n)
#        equ_area = shape.circle_equ(length,maxdiameter,width,pitch,total,half)
    elif device_shape == 1:
        edge = shape.squire(pitch*total,width,pitch,total,half)
#        iso = shape.squireiso(length,width,n)
#        equ_area = shape.squire_equ(length,maxdiameter,width,pitch,total,half)
#    single_cell.add(iso[0])
#    single_cell.add(iso[1])    
#  
#    #绘制外围等效曝光区
#    for i in equ_area:
#        for j in i:
#            single_cell.add(gdspy.Rectangle(*j,layer = 1,datatype = 1))
# 
        
    #纳米线核心区域点阵
    startpoint = edge[0]
    finalpoint = edge[1]
        
    i = 0        
    location = numpy.linspace(0,total,int(total/n)+1)   #隔离线位置
    j = 1.0  #direction
    max_length = 0
    while i< len(startpoint)-n*2:   #调整并联单元长度相等
        m = 0
        while m<n and i<len(startpoint):
            if i==0:    #保证0点的线存在
                startpoint[i]=(startpoint[i-m+1][0],startpoint[i][1])
                finalpoint[i]=(finalpoint[i-m+1][0],finalpoint[i][1])
            startpoint[i]=(startpoint[i-m][0],startpoint[i][1])
            finalpoint[i]=(finalpoint[i-m][0],finalpoint[i][1])
            m = m+1
            i = i+1
    i=0
    while i< len(startpoint):   #计算最大圆直径       
        if is_interrupt==1 and n>1:
                step = pals_length   #
                single_length = abs(finalpoint[i][0]-startpoint[i][0])
                if (int(single_length/step)+1)*step-length<n*pitch:
                   single_length = int(single_length/step+1)*step
                else:
                   single_length = int(single_length/step)*step
                if numpy.sqrt(single_length**2/4+startpoint[i][1]**2)>max_length and device_shape==0:  #圆形的隔离圆大小需要修改
                    max_length = numpy.sqrt((single_length**2/4+startpoint[i][1]**2))
                startpoint[i] = (-single_length/2,startpoint[i][1])
                finalpoint[i] = (single_length/2,finalpoint[i][1])
        i = i+1
    max_length= 2*max_length
    if max_length<pitch*total:
        max_length = pitch*total    #保证外切圆直径为离散值后的最大直径
    maxdiameter = max_length+2*minedge  #调整最大曝光区域
    if maxdiameter <= pin_length:
        maxdiameter = pin_length
    
    i = 0
    while i < len(startpoint):    #i=0，和total对应在x=0位置 
        if 0==1:
            j = -j            
        else:     #每隔30um绘制一个断点
            if is_interrupt==1 and n>1:
                step = pals_length   #
                k = 0
                if i not in location:
                    while (k+1)*step<finalpoint[i][0]-startpoint[i][0]:
                        steppoint1 = (startpoint[i][0]+k*step,startpoint[i][1])
                        steppoint2 = (startpoint[i][0]+(k+1)*step-width*(n-1),finalpoint[i][1])
                        polym = gdspy.Rectangle(steppoint1,steppoint2,layer = 1,datatype = 1)
                        polym.fillet(pitch-width,points_per_2pi = 20)
                        single_cell.add(polym)
                        k = k+1
                    else:
                        steppoint1 = (startpoint[i][0]+k*step,startpoint[i][1])
                        steppoint2 = finalpoint[i]
                        polym = gdspy.Rectangle(steppoint1,steppoint2,layer = 1,datatype = 1)
                        polym.fillet(pitch-width,points_per_2pi = 20)
                        single_cell.add(polym)
            else:
                if i not in location:
                    polym = gdspy.Rectangle(startpoint[i],
                                finalpoint[i],
                                layer = 1,datatype = 1)
                    polym.fillet((pitch-width)/2,points_per_2pi = 20)
                    single_cell.add(polym)
        if i in location:
            startpoint[i] = (j*startpoint[i][0],startpoint[i][1])
            if device_shape==0:
                finalpoint[i] = (j*numpy.sqrt((max_length/2+3)**2-((half-i)*pitch)**2),
                                 finalpoint[i][1])
            elif device_shape==1:
                finalpoint[i] = (j * (max_length / 2 + 3),
                                 finalpoint[i][1])
            j = -j
            polym = gdspy.Rectangle(startpoint[i],
                                finalpoint[i],
                                layer = 1,datatype = 1)
            polym.fillet((pitch-width)/2,points_per_2pi = 20)
            single_cell.add(polym)            
        i = i+1   
    if device_shape==0:
        iso = shape.circleiso(max_length,width,n)
        equ_area = shape.circle_equ(max_length,maxdiameter,width,pitch,total,half)
    elif device_shape == 1:
        iso = shape.squireiso(max_length,width,n)
        equ_area = shape.squire_equ(max_length,maxdiameter,width,pitch,total,half)
    single_cell.add(iso[0])
    single_cell.add(iso[1])    
  
    #绘制外围等效曝光区

    layer = {'layer':1,'datatype':1}  #定义图层
    for i in equ_area:
        for j in i:
            single_cell.add(gdspy.Rectangle(*j,**layer))

 
        
    #纳米线核心区域点阵
    startpoint = edge[0]
    finalpoint = edge[1]
    """
    电极线结构
    pitch 纳米线间距，即周期
    width:纳米线宽度
    length 纳米线长度，电极宽度设置为2um，在并联纳米线总宽宽度>1μm后，
    电极宽度调整为（n+4）*width
    预曝光区域，宽度设置为5μm
    曝光长度设置为length/2+5，pitch与纳米线一致
    """
    length = max_length   #将length改为划出纳米线调整后的最大直径
    total = int((maxdiameter/2-length/2-1)/pitch);
    i = 0
    if n*width >=1:
        half_elecwidth = (n+4)*width
    else:
        half_elecwidth = 1
    old = half_elecwidth
    
    while i <= total-1:
        if i<=int(n*m*width/pitch)+1:
            half_elecwidth = max_length/2+3
        else:
            half_elecwidth = old
        elec1 = gdspy.Rectangle((half_elecwidth,i*pitch+length/2),
                                (maxdiameter/2,i*pitch+length/2+pitch-width),**layer)
        single_cell.add(elec1)
        elec2 = gdspy.Rectangle((-half_elecwidth,i*pitch+length/2),
                                (-maxdiameter/2,i*pitch+length/2+pitch-width),**layer)
        single_cell.add(elec2)
        elec3 = gdspy.Rectangle((-half_elecwidth,(-i-1)*pitch-length/2),
                                (-maxdiameter/2,-i*pitch-length/2-(pitch-width)),**layer)
        single_cell.add(elec3)
        elec4 = gdspy.Rectangle((half_elecwidth,(-i-1)*pitch-length/2),
                                (maxdiameter/2,-i*pitch-length/2-(pitch-width)),**layer)
        single_cell.add(elec4)
        i = i+1


    """
    隔离线，宽度1um，伸出纳米线区域elec_width/2+pin_ength
    """
    length = oldl   #需要原始直径值
    if elec_width/2>length/2+pin_length-10:
        a = (elec_width/2-length/2-pin_length)/pin_length
        if a-int(a)>0.9:
            a = int(a)+2
        else:
            a = int(a)+1
        halflen = pin_length*(a+1)+length/2
    else:
        halflen = length/2+pin_length
    edgelen = maxdiameter/2 #纳米线区域边界
    insulate1 = gdspy.Rectangle((half_elecwidth,edgelen-1),
                                (halflen,edgelen),
                                **layer)
    single_cell.add(insulate1)
    insulate2 = gdspy.Rectangle((-half_elecwidth,edgelen-1),
                                (-halflen,edgelen),
                                **layer)
    single_cell.add(insulate2)
    insulate3 = gdspy.Rectangle((-half_elecwidth,-edgelen+1),
                                (-halflen,-edgelen),
                                **layer)
    single_cell.add(insulate3)
    insulate4 = gdspy.Rectangle((half_elecwidth,-edgelen+1),
                                (halflen,-edgelen),
                                **layer)
    single_cell.add(insulate4)
    
    insulate5 = gdspy.Rectangle((edgelen-1,halflen),
                                (edgelen,-halflen),
                                **layer)
    single_cell.add(insulate5)
    insulate6 = gdspy.Rectangle((-edgelen+1,halflen),
                                (-edgelen,-halflen),
                                **layer)
    single_cell.add(insulate6)

    """
    绘制拼接场外围尺寸，固定等效曝光区域宽10μm
    位置在边界处:pin_length-minedge
    """
    """
    length = oldl  # oldl=原始length，保证被整场分割而不滑移。允许出现边界处拼接场存在。
    layer_useless = {'layer': 3, 'datatype': 3}
    out = gdspy.Rectangle((-length / 2 - pin_length, -maxdiameter / 2),  # 左边
                          (-length / 2 - pin_length + 1, maxdiameter / 2),
                          **layer_useless
                          )
    single_cell.add(out)
    out = gdspy.Rectangle((length / 2 + pin_length, -maxdiameter / 2),  # 右边
                          (length / 2 + pin_length - 1, maxdiameter / 2),
                          **layer_useless
                          )
    single_cell.add(out)
    out = gdspy.Rectangle((-maxdiameter / 2, length / 2 + pin_length),  # 上左
                          (-half_elecwidth, length / 2 + pin_length - 1),
                          **layer_useless
                          )
    single_cell.add(out)
    out = gdspy.Rectangle((maxdiameter / 2, length / 2 + pin_length),  # 上右
                          (half_elecwidth, length / 2 + pin_length - 1),
                          **layer_useless
                          )
    single_cell.add(out)
    out = gdspy.Rectangle((-maxdiameter / 2, -length / 2 - pin_length),  # 下左
                          (-half_elecwidth, -length / 2 - pin_length + 1),
                          **layer_useless
                          )
    single_cell.add(out)
    out = gdspy.Rectangle((maxdiameter / 2, -length / 2 - pin_length),  # 下右
                          (half_elecwidth, -length / 2 - pin_length + 1),
                          **layer_useless
                          )
    single_cell.add(out)
    """

    """
    对准标记
    4umX100μm大小十字
    位置，-500um，500um
    """
    markers = [1,11,61,111,121]
    if mark in markers:
        cross = gdspy.Rectangle((-450,498),(-550,502),
                            **layer)
        single_cell.add(cross)
        cross = gdspy.Rectangle((-498,450),(-502,550),
                            **layer)
        single_cell.add(cross)
    
    
    """
    编号
    """
    text = gdspy.Text(str(number),20,(1000,300),
                      **layer)
    single_cell.add(text)
    
    return single_cell

def nanowire_array(dx,dy,m,n,cells):
    """
    绘制纳米线阵列
    """

    arrayname = 'array{}X{}'.format(m,n)
    refcell = gdspy.Cell(arrayname,exclude_from_current=True)
    lib = gdspy.GdsLibrary(name='lib')
#    lib.add(cells,overwrite_duplicate=True)   #2017102514：25
    j = m-1
    k = n
    for i in cells:
        refcell.add(gdspy.CellReference(i,rotation = 180).translate(dx*(k-1),dy*j))
        k = k-1
        if k%(n) ==0:
            k = n
            j = j-1
    refcell.flatten()      #
    lib.add(refcell,overwrite_duplicate=True)
    return lib

def get_param_from_xlsx(filename):
    """
    从xlsx文件中读取纳米线参数设定
    param1[[],]  单独纳米线结构参数
    param2 :[]   纳米线阵列参数
    """
    wb = opx.load_workbook(filename)
    sheetnames = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(sheetnames[0])
    param1 = []
    for rown in range(2,sheet.max_row+1):
        param1.append([])
        for coln in range(1,sheet.max_column+1):
            param1[rown-2].append(sheet.cell(row = rown,column = coln).value)
        param1[rown-2].append(rown-1)
    param2 = []
    sheet = wb.get_sheet_by_name(sheetnames[1])
    for coln in range(1,sheet.max_column+1):
        param2.append(sheet.cell(row = 2,column = coln).value)
    wb.close()
    print(param1)

    return param1,param2

def plot(filename):
    param1,param2 = get_param_from_xlsx(filename)
    if param2.count==0:
        param2 = [3250,3250,11,11,60,10]
    setting = {'elec_width':param2[-1],'pin_length':param2[-2]}
    cells=[nanowire(*i,**setting) for i in param1]
    param2.pop()
    param2.pop()
    param2.append(cells)
    lib = nanowire_array(*param2)
    dest = str(filename.replace('.xlsx','.gds'))
    lib.write_gds(dest,unit = 1e-06,precision=1e-09)
    return

def nTron(width,L,gatewidth):
    '''
    画nTron版图
    width 最窄地方宽度
    L 大小    single_cell = gdspy.Cell(Cellname,exclude_from_current=True)
    '''

